"""
Document Processor — parses PDF, DOCX, TXT/MD, XLSX, CSV files,
chunks the text, generates NVIDIA embeddings (pgvector), and stores
in Supabase for semantic + full-text search fallback.
All processing happens in-memory (no disk writes — Vercel compatible).
"""
import io
import os
import re
from models.db import get_client


# ── Embeddings ───────────────────────────────────────────────────────────────

def generate_embeddings(texts: list[str]) -> list[list[float] | None]:
    """
    Generate 1024-dim embeddings via NVIDIA NIM (nvidia/nv-embedqa-e5-v5).
    Batches up to 50 texts per call. Returns None per item on error so upload
    still succeeds (FTS fallback remains available).
    """
    from openai import OpenAI
    client = OpenAI(
        base_url=os.environ.get("NVIDIA_BASE_URL", "https://integrate.api.nvidia.com/v1"),
        api_key=os.environ.get("NVIDIA_API_KEY", ""),
    )
    results: list[list[float] | None] = []
    for i in range(0, len(texts), 50):
        batch = texts[i:i + 50]
        try:
            resp = client.embeddings.create(
                model="nvidia/nv-embedqa-e5-v5",
                input=batch,
                encoding_format="float",
                extra_body={"input_type": "passage", "truncate": "END"},
            )
            ordered = sorted(resp.data, key=lambda x: x.index)
            results.extend([item.embedding for item in ordered])
        except Exception:
            results.extend([None] * len(batch))
    return results


def query_embedding(text: str) -> list[float] | None:
    """Generate a query-side embedding for semantic search."""
    from openai import OpenAI
    client = OpenAI(
        base_url=os.environ.get("NVIDIA_BASE_URL", "https://integrate.api.nvidia.com/v1"),
        api_key=os.environ.get("NVIDIA_API_KEY", ""),
    )
    try:
        resp = client.embeddings.create(
            model="nvidia/nv-embedqa-e5-v5",
            input=[text],
            encoding_format="float",
            extra_body={"input_type": "query", "truncate": "END"},
        )
        return resp.data[0].embedding
    except Exception:
        return None


# ── Text extraction ──────────────────────────────────────────────────────────

def extract_text_pdf(file_bytes: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    return "\n\n".join(pages)


def extract_text_docx(file_bytes: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def extract_text_plain(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8", errors="replace")


def extract_text_xlsx(file_bytes: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else f"Column{i+1}"
                   for i, h in enumerate(rows[0])]
        lines = [f"## {sheet_name}"]
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            parts = []
            for header, cell in zip(headers, row):
                if cell is not None and str(cell).strip() != "":
                    parts.append(f"{header}: {cell}")
            if parts:
                lines.append(", ".join(parts) + ".")
        if len(lines) > 1:
            sections.append("\n".join(lines))
    wb.close()
    return "\n\n".join(sections)


def extract_text_csv(file_bytes: bytes) -> str:
    import csv
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    lines = []
    headers = reader.fieldnames or []
    if headers:
        lines.append(f"## Data ({', '.join(headers)})")
    for row in reader:
        parts = [f"{k}: {v}" for k, v in row.items() if v and str(v).strip()]
        if parts:
            lines.append(", ".join(parts) + ".")
    return "\n".join(lines)


def extract_text(file_bytes: bytes, file_type: str) -> str:
    ft = file_type.lower()
    if ft == "pdf":
        return extract_text_pdf(file_bytes)
    elif ft == "docx":
        return extract_text_docx(file_bytes)
    elif ft in ("xlsx", "xls"):
        return extract_text_xlsx(file_bytes)
    elif ft == "csv":
        return extract_text_csv(file_bytes)
    else:
        return extract_text_plain(file_bytes)


# ── Chunking ─────────────────────────────────────────────────────────────────

def chunk_text(text: str, chunk_size: int = 350, overlap: int = 100) -> list[str]:
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r" {2,}", " ", text)
    header_pattern = re.compile(r'^(#{1,3} .+|[A-Z][A-Z &/\-]{4,})$', re.MULTILINE)
    lines = text.splitlines()
    current_header = ""
    annotated_words = []
    for line in lines:
        stripped = line.strip()
        if header_pattern.match(stripped) and len(stripped) < 80:
            current_header = stripped.lstrip('#').strip()
        for word in stripped.split():
            annotated_words.append((word, current_header))
    chunks = []
    start = 0
    while start < len(annotated_words):
        end = min(start + chunk_size, len(annotated_words))
        segment_words  = [w for w, _ in annotated_words[start:end]]
        segment_header = annotated_words[start][1]
        chunk_body = " ".join(segment_words)
        if len(chunk_body.strip()) > 40:
            if segment_header and not chunk_body.startswith(segment_header):
                chunk = f"[{segment_header}]\n{chunk_body}"
            else:
                chunk = chunk_body
            chunks.append(chunk)
        start += chunk_size - overlap
    return chunks


# ── Supabase storage ─────────────────────────────────────────────────────────

def store_document(name: str, file_type: str, size_bytes: int) -> int:
    client = get_client()
    resp = client.table("documents").insert({
        "name":       name,
        "file_type":  file_type,
        "size_bytes": size_bytes,
    }).execute()
    return resp.data[0]["id"]


def store_chunks(document_id: int, document_name: str, chunks: list[str]):
    """
    Bulk-insert chunks with embeddings into doc_chunks.
    Generates embeddings in batches; falls back gracefully if API fails.
    """
    client = get_client()

    # Generate embeddings for all chunks
    embeddings = generate_embeddings(chunks)

    rows = []
    for i, (chunk, emb) in enumerate(zip(chunks, embeddings)):
        row = {
            "document_id":   document_id,
            "document_name": document_name,
            "chunk_index":   i,
            "content":       chunk,
        }
        if emb is not None:
            row["embedding"] = emb
        rows.append(row)

    # Insert in batches of 50 (embeddings are large)
    for i in range(0, len(rows), 50):
        client.table("doc_chunks").insert(rows[i:i + 50]).execute()


def delete_document(document_id: int):
    client = get_client()
    client.table("documents").delete().eq("id", document_id).execute()


# ── Main entry point ─────────────────────────────────────────────────────────

def process_and_store(filename: str, file_bytes: bytes, file_type: str) -> dict:
    """
    Full pipeline: extract → chunk → embed → store.
    Returns {"document_id": int, "chunks": int, "name": str}
    """
    text = extract_text(file_bytes, file_type)
    if not text.strip():
        raise ValueError("Could not extract any text from the document.")

    chunks = chunk_text(text)
    if not chunks:
        raise ValueError("Document produced no usable text chunks.")

    doc_id = store_document(filename, file_type, len(file_bytes))
    store_chunks(doc_id, filename, chunks)

    return {"document_id": doc_id, "chunks": len(chunks), "name": filename}


def get_all_documents() -> list[dict]:
    client = get_client()
    docs = (
        client.table("documents")
        .select("*")
        .order("created_at", desc=True)
        .execute()
        .data or []
    )
    for doc in docs:
        count_resp = (
            client.table("doc_chunks")
            .select("id", count="exact")
            .eq("document_id", doc["id"])
            .execute()
        )
        doc["chunk_count"] = count_resp.count or 0
    return docs
